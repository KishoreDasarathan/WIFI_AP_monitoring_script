import subprocess
import sys

# Function to install packages
def install_packages():
    packages = [
        "requests",
        "pandas",
        "colorama",
        "openpyxl",
        "urllib3"
    ]
    
    for package in packages:
        try:
            __import__(package)
        except ImportError:
            print(f"Installing {package}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# Call the function to install required packages
install_packages()

# Importing packages after installation
import requests
import pandas as pd
from datetime import datetime
import time
import urllib3
import os
from colorama import init, Fore
from openpyxl import load_workbook

# Initialize colorama
init(autoreset=True)

# Suppress InsecureRequestWarning
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# UniFi Controller settings
controller_url = 'https://enter your web controller ip'
username = 'enter your username'
password = 'enter your password'
site_id = 'default'  # Change if you're using a different site

# Excel file settings
excel_file = "Device_Activity.xlsx"

# Function to login to UniFi Controller
def login():
    session = requests.Session()
    login_url = f"{controller_url}/api/login"
    response = session.post(login_url, json={"username": username, "password": password}, verify=False)

    if response.status_code != 200:
        raise Exception("Failed to log in to UniFi Controller")

    return session

# Function to get connected clients for each AP
def get_connected_clients(session):
    clients_url = f"{controller_url}/api/s/{site_id}/stat/sta"
    response = session.get(clients_url, verify=False)

    if response.status_code != 200:
        raise Exception("Failed to retrieve clients")

    return response.json().get('data', [])  # Use .get to avoid KeyError

# Function to get AP details to map MAC addresses to names
def get_access_points(session):
    aps_url = f"{controller_url}/api/s/{site_id}/stat/device"
    response = session.get(aps_url, verify=False)

    if response.status_code != 200:
        raise Exception("Failed to retrieve access points")
    
    ap_data = response.json().get('data', [])
    return {ap['mac']: ap['name'] for ap in ap_data if 'mac' in ap and 'name' in ap}  # Map MAC addresses to names

# Function to convert bytes to human-readable format
def format_data_usage(bytes):
    if bytes < 1024:
        return f"{bytes} B"
    elif bytes < 1024**2:
        return f"{bytes / 1024:.2f} KB"
    elif bytes < 1024**3:
        return f"{bytes / (1024**2):.2f} MB"
    else:
        return f"{bytes / (1024**3):.2f} GB"

# Function to save activity to Excel
def save_to_excel(data, sheet_name):
    df = pd.DataFrame(data)

    # Try to load the existing workbook
    if os.path.exists(excel_file):
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Load existing workbook and check if the sheet exists
            workbook = load_workbook(excel_file)
            if sheet_name in workbook.sheetnames:
                startrow = workbook[sheet_name].max_row  # Find the next available row
                df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False)
            else:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

# Main function to monitor devices
def main():
    print("=== Monitoring Devices on UniFi AP ===")
    
    session = login()
    
    # Dictionary to track device connection state
    connected_devices = {}
    
    # Fetch AP names
    ap_names = get_access_points(session)

    # Initialize data storage for Excel
    log_data = []

    while True:
        clients = get_connected_clients(session)

        # Current connected devices
        current_devices = {client.get('mac'): client for client in clients if 'mac' in client}

        # Check for new connections
        for mac, client in current_devices.items():
            if mac not in connected_devices:
                # New connection
                connected_devices[mac] = {
                    'start_time': datetime.now(),
                    'data_usage': client.get('rx_bytes', 0) + client.get('tx_bytes', 0),  # Total data usage
                    'device_name': client.get('hostname', 'Unknown Device'),
                    'ap_mac': client.get('ap_mac', 'Unknown AP')
                }
                # Print device connection indication
                ap_location = ap_names.get(connected_devices[mac]['ap_mac'], 'Unknown Location')  # Get the location name from the map
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                print(f"\n{Fore.RED}[CONNECTED] [{current_time}] Device: {connected_devices[mac]['device_name']}\n  MAC: {mac}\n  AP Location: {ap_location}{Fore.RESET}\n")

                # Log to Excel
                log_data.append({
                    'Timestamp': current_time,
                    'MAC Address': mac,
                    'Action': 'Connected',
                    'Duration (s)': 0,
                    'Data Usage (bytes)': connected_devices[mac]['data_usage'],
                    'AP Location': ap_location
                })

        # Check for disconnections
        for mac in list(connected_devices.keys()):
            if mac not in current_devices:
                # Device disconnected
                start_time = connected_devices[mac]['start_time']
                duration = (datetime.now() - start_time).total_seconds()
                data_usage = connected_devices[mac]['data_usage']
                device_name = connected_devices[mac]['device_name']
                ap_mac = connected_devices[mac]['ap_mac']
                ap_location = ap_names.get(ap_mac, 'Unknown Location')
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                formatted_data_usage = format_data_usage(data_usage)

                # Print device disconnection indication
                print(f"\n{Fore.GREEN}[DISCONNECTED] [{current_time}] Device: {device_name}\n  MAC: {mac}\n  Duration: {duration:.2f} seconds\n  Data Usage: {formatted_data_usage}\n  AP Location: {ap_location}{Fore.RESET}\n")

                # Log to Excel
                log_data.append({
                    'Timestamp': current_time,
                    'MAC Address': mac,
                    'Action': 'Disconnected',
                    'Duration (s)': duration,
                    'Data Usage (bytes)': data_usage,
                    'AP Location': ap_location
                })

                # Remove device from tracking
                del connected_devices[mac]

                # Save log data to Excel after each disconnection
                sheet_name = datetime.now().strftime("%d-%m-%y")  # Date format as DD-MM-YY
                save_to_excel(log_data, sheet_name)
                log_data.clear()  # Clear the log data after saving

        # Wait before the next check
        time.sleep(10)  # Check every 10 seconds

if __name__ == "__main__":
    main()